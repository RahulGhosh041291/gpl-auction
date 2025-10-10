from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from database import get_db
from models import OwnerRegistration
import schemas
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

router = APIRouter()

@router.post("/owner-registrations/", response_model=schemas.OwnerRegistrationResponse)
async def create_owner_registration(
    registration: schemas.OwnerRegistrationCreate,
    db: Session = Depends(get_db)
):
    """Create a new owner registration"""
    db_registration = OwnerRegistration(
        owner_full_name=registration.owner_full_name,
        co_owner_full_name=registration.co_owner_full_name,
        owner_block=registration.owner_block,
        owner_unit_number=registration.owner_unit_number,
        co_owner_block=registration.co_owner_block,
        co_owner_unit_number=registration.co_owner_unit_number,
        interested_to_buy=registration.interested_to_buy,
        team_price=15000.0
    )
    
    db.add(db_registration)
    db.commit()
    db.refresh(db_registration)
    
    return db_registration

@router.get("/owner-registrations/", response_model=List[schemas.OwnerRegistrationResponse])
async def get_owner_registrations(db: Session = Depends(get_db)):
    """Get all owner registrations"""
    registrations = db.query(OwnerRegistration).order_by(OwnerRegistration.created_at.desc()).all()
    return registrations

@router.get("/owner-registrations/{registration_id}", response_model=schemas.OwnerRegistrationResponse)
async def get_owner_registration(registration_id: int, db: Session = Depends(get_db)):
    """Get a specific owner registration by ID"""
    registration = db.query(OwnerRegistration).filter(OwnerRegistration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")
    return registration

@router.get("/owner-registrations/export/excel")
async def export_owner_registrations_excel(db: Session = Depends(get_db)):
    """Export all owner registrations to Excel"""
    registrations = db.query(OwnerRegistration).order_by(OwnerRegistration.created_at.desc()).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Owner Registrations"
    
    # Define header style
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add event information at the top
    ws.merge_cells('A1:I1')
    event_cell = ws['A1']
    event_cell.value = "Galaxia Premier League Season 2 - Owner Registrations"
    event_cell.font = Font(bold=True, size=16, color="4F81BD")
    event_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A2:I2')
    event_details = ws['A2']
    event_details.value = "Tournament: 10-11 January 2026 | Auction: 2nd November 2025 at Club Stella | Team Price: ₹15,000"
    event_details.font = Font(size=11, italic=True)
    event_details.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add empty row
    ws.append([])
    
    # Headers
    headers = [
        "ID",
        "Owner Full Name",
        "Owner Block",
        "Owner Unit",
        "Co-Owner Full Name",
        "Co-Owner Block",
        "Co-Owner Unit",
        "Interested to Buy",
        "Team Price (INR)",
        "Registration Date"
    ]
    
    ws.append(headers)
    
    # Style headers
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data
    for reg in registrations:
        ws.append([
            reg.id,
            reg.owner_full_name,
            reg.owner_block.value,
            reg.owner_unit_number,
            reg.co_owner_full_name,
            reg.co_owner_block.value,
            reg.co_owner_unit_number,
            "Yes" if reg.interested_to_buy else "No",
            f"₹{reg.team_price:,.0f}",
            reg.created_at.strftime("%d-%b-%Y %I:%M %p")
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary section
    last_row = ws.max_row + 2
    ws.merge_cells(f'A{last_row}:B{last_row}')
    summary_cell = ws[f'A{last_row}']
    summary_cell.value = "Total Registrations:"
    summary_cell.font = Font(bold=True, size=11)
    summary_cell.alignment = Alignment(horizontal="right")
    ws[f'C{last_row}'] = len(registrations)
    ws[f'C{last_row}'].font = Font(bold=True, size=11)
    
    interested_count = sum(1 for reg in registrations if reg.interested_to_buy)
    last_row += 1
    ws.merge_cells(f'A{last_row}:B{last_row}')
    interested_cell = ws[f'A{last_row}']
    interested_cell.value = "Interested to Buy:"
    interested_cell.font = Font(bold=True, size=11, color="00B050")
    interested_cell.alignment = Alignment(horizontal="right")
    ws[f'C{last_row}'] = interested_count
    ws[f'C{last_row}'].font = Font(bold=True, size=11, color="00B050")
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename with timestamp
    filename = f"GPL_Season2_Owner_Registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.delete("/owner-registrations/{registration_id}")
async def delete_owner_registration(registration_id: int, db: Session = Depends(get_db)):
    """Delete an owner registration"""
    registration = db.query(OwnerRegistration).filter(OwnerRegistration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")
    
    db.delete(registration)
    db.commit()
    
    return {"message": "Registration deleted successfully"}
